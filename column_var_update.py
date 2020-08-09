#!/usr/bin/env python
# coding: utf-8

# ### Code by Hailey Tucker
# #### Last updated: 08 June, 2020


# Importing Libraries 
import string
import sys
import openpyxl
import xlsxwriter
import pandas as pd

###### Reading in the file and turning to a dataframe ###### 

# dynamically reading in system line- used for calling python script in Jenkins testing pipeline built with Perl
# sys_string = sys.argv 
# print ("This is the system string passed:", sys_string)
# input_file = ''
# input_file = str(sys_string[1])
# output_file = str(sys_string[2])
# target_column = str(sys_string[3])
# target_value = str(sys_string[4])

# hard-coding variables
input_file = 'manual_entries_temp.xlsx'
output_file = 'manual_entries_temp_new.xlsx'
target_column = 'POSTING_AMT'
target_value = 150 

# Reading in the file 
theFile = openpyxl.load_workbook(input_file)

# from the active attribute 
sheet_obj = theFile.active

# get max column count
max_column=sheet_obj.max_column
max_row=sheet_obj.max_row

excelDF = pd.ExcelFile(input_file)
df = pd.read_excel(excelDF, 'Sheet1')

df

###### Updating the correct cells #######
df.columns = df.columns.str.replace(' ', '')  # getting rid of extra white space on the side
df[target_column] = target_value
df


###### Taking those values and writing to a new excel sheet ########

df.to_excel(output_file, index = False)

